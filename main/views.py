from django.shortcuts import render
import openpyxl


def dropdown_view(request):
    workbook = openpyxl.load_workbook('src\kato.xlsx')
    sheet = workbook.active
    names = [row[0].value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1)]

    context = {'names': names}
    return render(request, 'cities.html', context)
